#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simple web interface for running the Rasch model (55 items) via a browser.
Runs locally on http://127.0.0.1:5000 by default.
"""

import os
import tempfile

import numpy as np
import pandas as pd
from flask import (
    Flask,
    render_template,
    request,
    send_file,
    redirect,
    url_for,
    flash,
)
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from scipy.optimize import minimize

from pdf2image import convert_from_path
from PIL import Image
from pyzbar.pyzbar import decode

import rash_manager_v2 as rm


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_12 = rm.TEMPLATE_12
STUDENTS_FILE = rm.STUDENTS_FILE
OUTPUT_DIR = rm.OUTPUT_DIR
KEYS_DIR = rm.KEYS_DIR
DPI = rm.DPI

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)
app.secret_key = "rash_web_secret_key"  # for flash messages


def run_rasch_on_dataframe(df: pd.DataFrame):
    """
    Core Rasch logic extracted from the desktop version.
    Expects a DataFrame with:
      - first column: student name
      - second column: ID (optional)
      - Q1..Q55 columns with 0/1 values
    Returns: (results_df, diff_df, warning_msg)
    """
    if df is None or df.empty:
        raise ValueError("Jadval bo'sh.")

    name_col = df.columns[0]
    id_col = df.columns[1] if len(df.columns) > 1 else None
    question_cols = [c for c in df.columns if str(c).upper().startswith("Q")]

    warning_msg = None
    if len(question_cols) != 55:
        warning_msg = f"{len(question_cols)} ta savol topildi. 55 bo'lishi kerak. Davom etiladi."

    names = df[name_col]
    ids = df[id_col] if id_col is not None else pd.Series([""] * len(df))
    answers = df[question_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)

    n_students, n_items = answers.shape
    if n_students == 0 or n_items == 0:
        raise ValueError("Savollar yoki o'quvchilar topilmadi.")

    # Weights: Q36–Q55 og'irroq
    weights = np.ones(n_items)
    if n_items > 35:
        weights[35:] = 1.5

    def neg_ll(params):
        theta = params[:n_students]
        beta = params[n_students:]
        P = 1 / (1 + np.exp(-(theta.reshape(-1, 1) - beta.reshape(1, -1))))
        P = np.clip(P, 1e-9, 1 - 1e-9)
        ll = np.sum(weights * (answers.values * np.log(P) + (1 - answers.values) * np.log(1 - P)))
        return -ll

    params0 = np.zeros(n_students + n_items)
    res = minimize(neg_ll, params0, method="L-BFGS-B", options={"maxiter": 500})
    params = res.x

    theta = params[:n_students]
    beta = params[n_students:]
    beta -= np.mean(beta)

    Z = (theta - np.mean(theta)) / np.std(theta)
    T = 50 + 10 * Z
    T = np.clip(T, 10, 90)

    def baho(t):
        if t >= 70:
            return "A+"
        if t >= 65:
            return "A"
        if t >= 60:
            return "B+"
        if t >= 55:
            return "B"
        if t >= 50:
            return "C+"
        if t >= 45:
            return "C"
        return "Sertifikat berilmadi"

    grades = [baho(t) for t in T]
    raw_scores = answers.sum(axis=1)

    results = pd.DataFrame(
        {
            "Ism Familiya": names,
            "ID": ids,
            "To‘g‘ri javoblar soni": raw_scores,
            "Qobiliyat (θ)": np.round(theta, 3),
            "Z ball": np.round(Z, 3),
            "T ball": np.round(T, 2),
            "Baholash": grades,
        }
    )
    # Sort by T descending
    results = results.sort_values(by="T ball", ascending=False)

    diff = pd.DataFrame(
        {
            "Savol": question_cols,
            "Qiyinlik (β)": np.round(beta, 3),
        }
    )

    return results, diff, warning_msg


def save_rasch_to_excel_with_formatting(results_df: pd.DataFrame, diff_df: pd.DataFrame):
    """
    Saves Rasch results + difficulty sheet into a temporary XLSX file,
    applies same colors and bar-chart as the desktop version, and returns the temp path.
    """
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()

    # First write data
    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="Natijalar", index=False)
        diff_df.to_excel(writer, sheet_name="Savol Qiyinliklari", index=False)

    # Then re-open with openpyxl to apply styles and chart
    wb = load_workbook(tmp_path)
    ws = wb["Natijalar"]
    ws_diff = wb["Savol Qiyinliklari"]

    colors = {
        "A+": "000080",
        "A": "3399FF",
        "B+": "00FF00",
        "B": "FFFF00",
        "C+": "FF9900",
        "C": "FFCC00",
        "Sertifikat berilmadi": "FF0000",
    }

    grade_col = list(results_df.columns).index("Baholash") + 1

    # Helper to convert column index -> Excel letter
    def col_letter(n: int) -> str:
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    grade_letter = col_letter(grade_col)

    for r in range(2, ws.max_row + 1):
        grade_value = ws[f"{grade_letter}{r}"].value
        col = colors.get(grade_value)
        if col:
            fill = PatternFill(start_color=col, end_color=col, fill_type="solid")
            for c in range(1, ws.max_column + 1):
                ws[f"{col_letter(c)}{r}"].fill = fill

    # Create bar chart (difficulty diagram)
    chart = BarChart()
    chart.title = "Savollar qiyinligi (β)"
    chart.x_axis.title = "Savollar"
    chart.y_axis.title = "Qiyinlik"

    data = Reference(ws_diff, min_col=2, min_row=1, max_row=ws_diff.max_row)
    cats = Reference(ws_diff, min_col=1, min_row=2, max_row=ws_diff.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_diff.add_chart(chart, "E2")

    wb.save(tmp_path)
    return tmp_path


@app.route("/", methods=["GET"])
def index():
    students = load_students()
    tests = list_tests()
    tests_info = [
        {
            "name": t,
            "has_tituls": test_has_tituls(t),
        }
        for t in tests
    ]
    return render_template(
        "index.html",
        students=students.to_dict(orient="records"),
        tests=tests_info,
    )


#########################
#  STUDENTS & TESTS
#########################


def load_students():
    """Web-safe wrapper around the desktop student storage."""
    if not os.path.exists(STUDENTS_FILE):
        pd.DataFrame(columns=["Ism Familiya", "ID"]).to_excel(STUDENTS_FILE, index=False)
    return pd.read_excel(STUDENTS_FILE)


def save_students(df: pd.DataFrame):
    df.to_excel(STUDENTS_FILE, index=False)


def list_tests():
    """List existing tests based on key files."""
    if not os.path.exists(KEYS_DIR):
        return []
    tests = []
    for name in os.listdir(KEYS_DIR):
        if name.lower().endswith(".xlsx"):
            tests.append(os.path.splitext(name)[0])
    tests.sort()
    return tests


def titul_dir_for_test(test_name: str) -> str:
    base, tituls, _ = rm.ensure_test_dirs(test_name)
    return tituls


def test_has_tituls(test_name: str) -> bool:
    tituls = titul_dir_for_test(test_name)
    if not os.path.exists(tituls):
        return False
    # any PDF inside tituls folder
    for f in os.listdir(tituls):
        if f.lower().endswith(".pdf"):
            return True
    return False


def load_pixel_points_for_web():
    """Same as rm.load_pixel_points but without GUI popups."""
    if not os.path.exists(rm.COORD_EXCEL):
        raise FileNotFoundError(f"Excel topilmadi: {rm.COORD_EXCEL}")
    df = pd.read_excel(rm.COORD_EXCEL)
    df.columns = [c.strip().lower() for c in df.columns]
    if not all(k in df.columns for k in ["savol", "x", "y"]):
        raise RuntimeError("Excelda ustunlar kerak: Savol, X, Y (Variant optional).")
    df["savol"] = df["savol"].astype(int)
    df["x"] = df["x"].astype(int)
    df["y"] = df["y"].astype(int)
    if "variant" in df.columns:
        df["variant"] = df["variant"].astype(str).str.strip().str.upper()
    else:
        df["variant"] = ""
    return df


def generate_tituls_for_test(test_name: str):
    """
    Web wrapper around rash_manager_v2.generate_tituls_for_test_name so
    that the generated titles (including QR size/placement) are identical
    between the desktop and web versions.
    """
    if not test_name:
        raise ValueError("Test nomi ko'rsatilmagan.")
    tituls_dir = rm.generate_tituls_for_test_name(test_name)
    return tituls_dir


#########################
#  GRADING HELPERS (WEB)
#########################


def _grade_front_image_pil(pil_img: Image.Image):
    """Core logic for front image (Q1..Q35) grading, reused by multiple endpoints."""
    coords = load_pixel_points_for_web()
    img_width, img_height = pil_img.size
    scaled_coords = rm.scale_coordinates(coords, img_width, img_height)
    front_coords = scaled_coords[scaled_coords["savol"] <= 35]
    scaled_params = rm.scale_detection_params(img_width, img_height)

    img_np = np.array(pil_img)[:, :, ::-1]
    codes = decode(pil_img)
    if not codes:
        raise RuntimeError("QR kod rasmda topilmadi.")

    import json

    try:
        data = json.loads(codes[0].data.decode("utf-8"))
    except Exception:
        raise RuntimeError("QR parse xato.")

    test_name = data.get("test")
    sid = str(data.get("id")).replace("TST", "")
    key = rm.load_key(test_name)
    if key is None:
        raise RuntimeError(f"{test_name} kaliti topilmadi.")

    detected = rm.detect_bubbles(
        img_np,
        front_coords,
        bubble_window=scaled_params["bubble_window"],
        min_black_pixels=scaled_params["min_black_pixels"],
    )

    students = load_students()
    row = {"Ism Familiya": "", "ID": sid, "Test": test_name}
    if sid in students["ID"].astype(str).tolist():
        row["Ism Familiya"] = students.loc[students["ID"].astype(str) == sid, "Ism Familiya"].values[0]

    for q in range(1, 36):
        d = detected.get(q)
        try:
            exp = str(key.loc[q - 1, "Javob"]).strip().upper()
        except Exception:
            exp = ""
        row[f"Q{q}"] = 1 if (d and exp and d == exp) else 0
    return row


def _grade_back_image_pil(pil_img: Image.Image):
    """Core logic for back image (Q36..Q55) grading."""
    coords = load_pixel_points_for_web()
    img_width, img_height = pil_img.size
    scaled_coords = rm.scale_coordinates(coords, img_width, img_height)
    back_coords = scaled_coords[(scaled_coords["savol"] >= 36) & (scaled_coords["savol"] <= 55)]
    if back_coords.empty:
        raise RuntimeError("Orqa uchun koordinatalar topilmadi.")

    scaled_params = rm.scale_detection_params(img_width, img_height)
    img_np = np.array(pil_img)[:, :, ::-1]
    results = rm.detect_back_squares(
        img_np,
        back_coords,
        square_half=scaled_params["back_square_half"],
        back_black_threshold=scaled_params["back_black_threshold"],
    )
    row = {}
    for q in range(36, 56):
        row[f"Q{q}"] = results.get(q, 0)
    return row


@app.route("/students/add", methods=["POST"])
def add_student():
    name = request.form.get("name", "").strip()
    sid = request.form.get("sid", "").strip()
    if not name or not sid:
        flash("Ism Familiya va ID majburiy.", "error")
        return redirect(url_for("index"))

    df = load_students()
    if str(sid) in df["ID"].astype(str).tolist():
        flash("Bu ID allaqachon mavjud.", "error")
        return redirect(url_for("index"))

    df = pd.concat([df, pd.DataFrame([{"Ism Familiya": name, "ID": sid}])], ignore_index=True)
    save_students(df)
    flash(f"{name} qo‘shildi.", "success")
    return redirect(url_for("index"))


@app.route("/students/upload", methods=["POST"])
def upload_students():
    file = request.files.get("students_file")
    if not file or file.filename == "":
        flash("O'quvchilar ro'yxati uchun Excel faylni tanlang.", "error")
        return redirect(url_for("index"))

    try:
        new_df = pd.read_excel(file)
    except Exception as e:
        flash(f"Excel faylni o'qishda xato: {e}", "error")
        return redirect(url_for("index"))

    # Expect at least these columns
    required_cols = {"Ism Familiya", "ID"}
    if not required_cols.issubset(set(new_df.columns)):
        flash("Excel faylida 'Ism Familiya' va 'ID' ustunlari bo'lishi kerak.", "error")
        return redirect(url_for("index"))

    # Normalize and merge, avoiding duplicate IDs
    new_df = new_df[list(required_cols)].copy()
    new_df["ID"] = new_df["ID"].astype(str)

    existing = load_students()
    existing["ID"] = existing["ID"].astype(str)

    before_count = len(existing)
    # Drop rows whose ID already exists
    merged = pd.concat(
        [
            existing,
            new_df[~new_df["ID"].isin(existing["ID"])],
        ],
        ignore_index=True,
    )
    added_count = len(merged) - before_count

    save_students(merged)
    flash(f"{added_count} ta yangi o'quvchi qo'shildi (jami: {len(merged)}).", "success")
    return redirect(url_for("index"))


@app.route("/tests/create", methods=["POST"])
def create_test():
    test_name = request.form.get("test_name", "").strip()
    if not test_name:
        flash("Test nomi majburiy.", "error")
        return redirect(url_for("index"))

    rm.ensure_test_dirs(test_name)
    rm.create_key_if_missing(test_name)
    flash(f"Test '{test_name}' yaratildi.", "success")
    return redirect(url_for("index"))


@app.route("/tests/upload_key", methods=["POST"])
def upload_test_key():
    test_name = request.form.get("key_test_name", "").strip()

    if not test_name:
        flash("Kalitni saqlash uchun testni tanlang.", "error")
        return redirect(url_for("index"))

    # Ensure directories exist for this test
    rm.ensure_test_dirs(test_name)

    try:
        # Read answers from form fields q1..q35 (front side only)
        savollar = []
        javoblar = []
        for i in range(1, 36):
            ans = (request.form.get(f"q{i}", "") or "").strip().upper()
            savollar.append(i)
            javoblar.append(ans)

        # At least one answer should be non-empty
        if all(a == "" for a in javoblar):
            flash("Hech bo'lmaganda bitta savol uchun javob kiriting.", "error")
            return redirect(url_for("index"))

        df = pd.DataFrame({"Savol": savollar, "Javob": javoblar})

        key_path = os.path.join(KEYS_DIR, f"{test_name}.xlsx")
        df.to_excel(key_path, index=False)
    except Exception as e:
        flash(f"Kalitni saqlashda xato: {e}", "error")
        return redirect(url_for("index"))

    flash(f"'{test_name}' testi uchun kalit saqlandi.", "success")
    return redirect(url_for("index"))


@app.route("/tituls/generate", methods=["POST"])
def tituls_generate():
    test_name = request.form.get("test_for_titul", "").strip()
    if not test_name:
        flash("Titul yaratish uchun testni tanlang.", "error")
        return redirect(url_for("index"))

    try:
        tituls_dir = generate_tituls_for_test(test_name)
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("index"))

    flash(f"Titullar yaratildi: {tituls_dir}", "success")
    return redirect(url_for("index"))


@app.route("/tituls/download", methods=["GET"])
def tituls_download():
    test_name = request.args.get("test")
    if not test_name:
        flash("Test nomi ko'rsatilmagan.", "error")
        return redirect(url_for("index"))

    tituls_dir = titul_dir_for_test(test_name)
    if not os.path.exists(tituls_dir):
        flash("Bu test uchun titul papkasi topilmadi.", "error")
        return redirect(url_for("index"))

    pdf_files = [f for f in os.listdir(tituls_dir) if f.lower().endswith(".pdf")]
    if not pdf_files:
        flash("Bu test uchun hech qanday titul PDF topilmadi. Avval titullarni yarating.", "error")
        return redirect(url_for("index"))

    import shutil

    tmp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(tmp_dir, f"{test_name}_tituls.zip")

    # Create zip with all PDFs in tituls folder
    shutil.make_archive(zip_path[:-4], "zip", tituls_dir)

    return send_file(
        zip_path,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"{test_name}_tituls.zip",
    )


@app.route("/run_rasch", methods=["POST"])
def run_rasch_endpoint():
    file = request.files.get("results_file")
    if not file or file.filename == "":
        flash("Iltimos, Excel faylni tanlang.", "error")
        return redirect(url_for("index"))

    try:
        df = pd.read_excel(file)
    except Exception as e:
        flash(f"Faylni o'qishda xato: {e}", "error")
        return redirect(url_for("index"))

    try:
        results_df, diff_df, warning_msg = run_rasch_on_dataframe(df)
        if warning_msg:
            flash(warning_msg, "warning")
    except Exception as e:
        flash(str(e), "error")
        return redirect(url_for("index"))

    tmp_path = save_rasch_to_excel_with_formatting(results_df, diff_df)

    # Let browser download the file
    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="rasch_55_weight_chart.xlsx",
    )


#########################
#  GRADING ROUTES (WEB)
#########################


@app.route("/grade/front_image", methods=["POST"])
def grade_front_image_route():
    file = request.files.get("front_image")
    if not file or file.filename == "":
        flash("Old tomonni tekshirish uchun rasm tanlang.", "error")
        return redirect(url_for("index"))
    try:
        pil = Image.open(file.stream).convert("RGB")
        row = _grade_front_image_pil(pil)
    except Exception as e:
        flash(f"Old tomonni tekshirishda xato: {e}", "error")
        return redirect(url_for("index"))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame([row]).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="front_image_result.xlsx",
    )


@app.route("/grade/front_images_batch", methods=["POST"])
def grade_front_images_batch_route():
    files = request.files.getlist("front_images")
    if not files:
        flash("Kamida bitta old tomon rasm tanlang.", "error")
        return redirect(url_for("index"))

    rows = []
    errors = []
    for f in files:
        if not f or f.filename == "":
            continue
        try:
            pil = Image.open(f.stream).convert("RGB")
            row = _grade_front_image_pil(pil)
            row["Fayl"] = f.filename
            rows.append(row)
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    if not rows:
        msg = "Hech qanday rasm muvaffaqiyatli tekshirilmadi."
        if errors:
            msg += " " + "; ".join(errors[:5])
        flash(msg, "error")
        return redirect(url_for("index"))

    if errors:
        flash(f"Ba'zi fayllarda xato: {len(errors)} ta.", "warning")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame(rows).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="front_images_batch_result.xlsx",
    )


@app.route("/grade/back_image", methods=["POST"])
def grade_back_image_route():
    file = request.files.get("back_image")
    if not file or file.filename == "":
        flash("Orqa tomonni tekshirish uchun rasm tanlang.", "error")
        return redirect(url_for("index"))
    try:
        pil = Image.open(file.stream).convert("RGB")
        row = _grade_back_image_pil(pil)
    except Exception as e:
        flash(f"Orqa tomonni tekshirishda xato: {e}", "error")
        return redirect(url_for("index"))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame([row]).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="back_image_result.xlsx",
    )


@app.route("/grade/back_images_batch", methods=["POST"])
def grade_back_images_batch_route():
    files = request.files.getlist("back_images")
    if not files:
        flash("Kamida bitta orqa tomon rasm tanlang.", "error")
        return redirect(url_for("index"))

    rows = []
    errors = []
    for f in files:
        if not f or f.filename == "":
            continue
        try:
            pil = Image.open(f.stream).convert("RGB")
            row = _grade_back_image_pil(pil)
            row["Fayl"] = f.filename
            rows.append(row)
        except Exception as e:
            errors.append(f"{f.filename}: {e}")

    if not rows:
        msg = "Hech qanday orqa rasm muvaffaqiyatli tekshirilmadi."
        if errors:
            msg += " " + "; ".join(errors[:5])
        flash(msg, "error")
        return redirect(url_for("index"))

    if errors:
        flash(f"Ba'zi orqa fayllarda xato: {len(errors)} ta.", "warning")

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame(rows).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="back_images_batch_result.xlsx",
    )


@app.route("/grade/back_pdf", methods=["POST"])
def grade_back_pdf_route():
    file = request.files.get("back_pdf")
    if not file or file.filename == "":
        flash("Orqa tomonni tekshirish uchun PDF tanlang.", "error")
        return redirect(url_for("index"))
    try:
        # Save to temp file for pdf2image
        tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        file.stream.seek(0)
        tmp_pdf.write(file.stream.read())
        tmp_pdf_path = tmp_pdf.name
        tmp_pdf.close()

        pages = convert_from_path(tmp_pdf_path, dpi=DPI)
        if len(pages) < 2:
            raise RuntimeError("PDF ichida orqa sahifa topilmadi.")
        pil = pages[1].convert("RGB")
        row = _grade_back_image_pil(pil)
    except Exception as e:
        flash(f"Orqa PDF sahifasini tekshirishda xato: {e}", "error")
        return redirect(url_for("index"))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame([row]).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="back_pdf_result.xlsx",
    )


@app.route("/grade/front_pdf_pages", methods=["POST"])
def grade_front_pdf_pages_route():
    file = request.files.get("front_pdf")
    if not file or file.filename == "":
        flash("Old tomonni tekshirish uchun PDF tanlang.", "error")
        return redirect(url_for("index"))

    rows = []
    try:
        tmp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        file.stream.seek(0)
        tmp_pdf.write(file.stream.read())
        tmp_pdf_path = tmp_pdf.name
        tmp_pdf.close()

        pages = convert_from_path(tmp_pdf_path, dpi=DPI)
        for i, page in enumerate(pages, start=1):
            try:
                pil = page.convert("RGB")
                row = _grade_front_image_pil(pil)
                row["Sahifa"] = i
                rows.append(row)
            except Exception as e:
                # Skip problematic pages but continue others
                rows.append({"Sahifa": i, "Xato": str(e)})
    except Exception as e:
        flash(f"Old PDF sahifalarini tekshirishda xato: {e}", "error")
        return redirect(url_for("index"))

    if not rows:
        flash("Hech qanday sahifa muvaffaqiyatli tekshirilmadi.", "error")
        return redirect(url_for("index"))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_path = tmp.name
    tmp.close()
    pd.DataFrame(rows).to_excel(tmp_path, index=False)

    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="front_pdf_pages_result.xlsx",
    )


if __name__ == "__main__":
    # Run on localhost:5000
    app.run(host="0.0.0.0", debug=True)


