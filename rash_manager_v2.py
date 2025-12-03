#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
rash_manager_v2.py — Full (front 1-35 bubbles, back 36-55 circles), JPG+PDF checks, DPI=300
Author: generated for Mirkomil
"""

import os
import json
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd
import numpy as np
from pdf2image import convert_from_path
from PIL import Image, ImageDraw, ImageFont
import qrcode
from pyzbar.pyzbar import decode
import cv2
from scipy.optimize import minimize
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference

# ------------- CONFIG -------------
# Determine BASE_DIR: use executable directory if running as exe, otherwise use script directory
if getattr(sys, 'frozen', False):
    # Running as compiled executable
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Running as script
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOADED_12 = "/mnt/data/12.pdf"
TEMPLATE_12 = UPLOADED_12 if os.path.exists(UPLOADED_12) else os.path.join(BASE_DIR, "Titul.pdf")

COORD_EXCEL = os.path.join(BASE_DIR, "titul_bubble_koordinatalar_2480x3508.xlsx")
STUDENTS_FILE = os.path.join(BASE_DIR, "students.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
KEYS_DIR = os.path.join(BASE_DIR, "test_keys")
KEYWORD_FLAG_FILE = os.path.join(BASE_DIR, ".keyword_verified")
KEYWORD = "Mirkomil12"

# QR placement on front template (these are defaults — if template different, adjust)
QR_X_DEFAULT = 1400
QR_Y_DEFAULT = 300
QR_SIZE = 550

# DPI for pdf2image conversion
DPI = 300

# Reference dimensions for coordinates in Excel file
REF_WIDTH = 2480
REF_HEIGHT = 3508

# Detection params tuned for 2480x3508 JPG front/back coordinates
# front bubble radius optimized for 2480x3508 scans
BUBBLE_RADIUS = 26
BUBBLE_WINDOW = 52  # 52×52 px
THRESH_MULT = 1.15  # belgilangan bubble kuchli ajralishi uchun (tuned for new size)
MIN_BLACK_PIXELS = 1400  # minimum absolute threshold for a marked bubble

# back square detection (painted square check)
BACK_SQUARE_SIDE = 91  # Square side length in pixels
BACK_SQUARE_HALF = BACK_SQUARE_SIDE // 2  # Half side for centering (45px)
BACK_BLACK_THRESHOLD = 350  # Threshold for painted square detection

os.makedirs(BASE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(KEYS_DIR, exist_ok=True)

# ------------- HELPERS -------------
def check_keyword_verification():
    """
    Check keyword verification. If not verified yet, ask for keyword 'Mirkomil12' 
    only on first run. Create flag file to remember verification for subsequent runs.
    Returns True if verified (or already verified), False if wrong keyword entered.
    """
    # Check if already verified
    if os.path.exists(KEYWORD_FLAG_FILE):
        return True
    
    # First run - ask for keyword
    root_check = tk.Tk()
    root_check.withdraw()  # Hide main window
    root_check.title("Kalit so'z")
    
    keyword_input = simpledialog.askstring(
        "Kalit so'z", 
        "Dasturni ishga tushirish uchun kalit so'zni kiriting:",
        show='*',
        parent=root_check
    )
    root_check.destroy()
    
    if keyword_input == KEYWORD:
        # Correct keyword - create flag file
        with open(KEYWORD_FLAG_FILE, 'w') as f:
            f.write('verified')
        return True
    else:
        # Wrong keyword
        messagebox.showerror("Xato", "Noto'g'ri kalit so'z! Dastur yopilmoqda.")
        return False

def ensure_test_dirs(test_name):
    base = os.path.join(OUTPUT_DIR, test_name)
    tituls = os.path.join(base, "tituls")
    checked = os.path.join(base, "checked")
    os.makedirs(tituls, exist_ok=True)
    os.makedirs(checked, exist_ok=True)
    return base, tituls, checked

def load_students():
    if not os.path.exists(STUDENTS_FILE):
        pd.DataFrame(columns=["Ism Familiya", "ID"]).to_excel(STUDENTS_FILE, index=False)
    return pd.read_excel(STUDENTS_FILE)

def save_students(df):
    df.to_excel(STUDENTS_FILE, index=False)

def create_key_if_missing(test_name):
    path = os.path.join(KEYS_DIR, f"{test_name}.xlsx")
    if not os.path.exists(path):
        df = pd.DataFrame({"Savol": list(range(1, 56)), "Javob": [""] * 55})
        df.to_excel(path, index=False)
    return path

def load_key(test_name):
    path = os.path.join(KEYS_DIR, f"{test_name}.xlsx")
    return pd.read_excel(path) if os.path.exists(path) else None

def load_pixel_points():
    if not os.path.exists(COORD_EXCEL):
        messagebox.showerror("Xato", f"Excel topilmadi: {COORD_EXCEL}")
        return None
    df = pd.read_excel(COORD_EXCEL)
    # normalize column names
    df.columns = [c.strip().lower() for c in df.columns]
    # require columns
    if not all(k in df.columns for k in ['savol', 'x', 'y']):
        messagebox.showerror("Xato", "Excelda ustunlar kerak: Savol, X, Y (Variant optional).")
        return None
    df['savol'] = df['savol'].astype(int)
    df['x'] = df['x'].astype(int)
    df['y'] = df['y'].astype(int)
    if 'variant' in df.columns:
        df['variant'] = df['variant'].astype(str).str.strip().str.upper()
    else:
        df['variant'] = ''
    return df

def scale_coordinates(coords_df, img_width, img_height, ref_width=REF_WIDTH, ref_height=REF_HEIGHT):
    """
    Scale coordinates from reference dimensions to actual image dimensions.
    
    Args:
        coords_df: DataFrame with columns ['savol', 'x', 'y', ...]
        img_width: Actual image width
        img_height: Actual image height
        ref_width: Reference width (default: 2480)
        ref_height: Reference height (default: 3508)
    
    Returns:
        DataFrame with scaled coordinates
    """
    df = coords_df.copy()
    scale_x = img_width / ref_width
    scale_y = img_height / ref_height
    df['x'] = (df['x'] * scale_x).astype(int)
    df['y'] = (df['y'] * scale_y).astype(int)
    return df

def scale_detection_params(img_width, img_height, ref_width=REF_WIDTH, ref_height=REF_HEIGHT):
    """
    Calculate scaled detection parameters based on image dimensions.
    
    Args:
        img_width: Actual image width
        img_height: Actual image height
        ref_width: Reference width (default: 2480)
        ref_height: Reference height (default: 3508)
    
    Returns:
        Dictionary with scaled parameters: {'bubble_window', 'back_radius', 'min_black_pixels', 'back_black_threshold'}
    """
    # Use average scale factor for detection parameters to maintain aspect ratio consistency
    scale_avg = ((img_width / ref_width) + (img_height / ref_height)) / 2
    return {
        'bubble_window': int(BUBBLE_WINDOW * scale_avg),
        'back_square_half': int(BACK_SQUARE_HALF * scale_avg),
        'min_black_pixels': int(MIN_BLACK_PIXELS * scale_avg * scale_avg),  # Area scales with scale^2
        'back_black_threshold': int(BACK_BLACK_THRESHOLD * scale_avg * scale_avg)  # Area scales with scale^2
    }

# ------------- TITUL GENERATION (front + back -> per-student 2-page PDF) -------------
def generate_tituls_for_test_name(test_name: str):
    """
    Core titul generation used by both the desktop GUI and the web app.
    This logic must stay in sync so that titles/QR sizes are identical.
    """
    if not test_name:
        return None
    if not os.path.exists(TEMPLATE_12):
        raise FileNotFoundError(f"12.pdf topilmadi: {TEMPLATE_12}")
    try:
        pages = convert_from_path(TEMPLATE_12, dpi=DPI)
    except Exception as e:
        raise RuntimeError(f"12.pdf ochishda xato: {e}")
    if len(pages) < 2:
        raise RuntimeError("12.pdf ichida kamida 2 sahifa (front + back) bo'lishi kerak.")
    front_template = pages[0].convert("RGB")
    back_template = pages[1].convert("RGB")
    students = load_students()
    if students.empty:
        raise RuntimeError("students.xlsx bo‘sh.")
    base, tituls, _ = ensure_test_dirs(test_name)
    # fonts
    try:
        font_small = ImageFont.truetype("Arial.ttf", 24)
        font_big = ImageFont.truetype("Arial.ttf", 36)
    except Exception:
        font_small = ImageFont.load_default()
        font_big = ImageFont.load_default()
    for _, row in students.iterrows():
        name = str(row["Ism Familiya"])
        sid = str(row["ID"])
        test_id = f"TST{sid}"
        payload = json.dumps({"test": test_name, "id": test_id})
        qr = qrcode.QRCode(
            version=2,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2,
        )
        qr.add_data(payload)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
        qr_big_front = qr_img.resize((QR_SIZE, QR_SIZE))
        # FRONT
        fr = front_template.copy()
        draw_f = ImageDraw.Draw(fr)
        w_f, h_f = fr.size
        # Slightly move QR up and to the right on the front side (same size)
        qx = min(QR_X_DEFAULT, max(0, w_f - QR_SIZE))  # a bit more to the right
        qy = min(max(QR_Y_DEFAULT, 0), max(0, h_f - QR_SIZE))  # a bit higher
        fr.paste(qr_big_front, (qx, qy))
        tx = qx + QR_SIZE + 30
        ty = qy + 10
        draw_f.text((tx, ty), f"Ism Familiya: {name}", fill="black", font=font_small)
        draw_f.text((tx, ty + 30), f"ID: {sid}", fill="black", font=font_small)
        draw_f.text((tx, ty + 60), f"Test: {test_name}", fill="black", font=font_small)
        # BACK: smaller QR, slightly higher, name+ID to right of QR
        bk = back_template.copy()
        draw_b = ImageDraw.Draw(bk)
        w_b, h_b = bk.size
        back_qr_size = int(QR_SIZE * 0.35)  # make back-side QR smaller
        qr_big_back = qr_img.resize((back_qr_size, back_qr_size))
        qr_x = max(10, w_b // 2 - back_qr_size // 2)
        qr_y = 20  # a little higher than before
        bk.paste(qr_big_back, (qr_x, qr_y))
        text_x = qr_x + back_qr_size + 40
        text_y = qr_y + 10
        draw_b.text((text_x, text_y), name, fill="black", font=font_big)
        draw_b.text((text_x, text_y + 45), f"ID: {sid}", fill="black", font=font_big)
        out_pdf = os.path.join(tituls, f"{sid}_{test_id}.pdf")
        fr.save(out_pdf, "PDF", save_all=True, append_images=[bk])
    create_key_if_missing(test_name)
    return tituls


def generate_tituls():
    """GUI wrapper that asks for test name then calls the core generator."""
    test_name = simpledialog.askstring("Titul yaratish", "Test nomini kiriting:")
    if not test_name:
        return
    try:
        tituls = generate_tituls_for_test_name(test_name)
    except Exception as e:
        messagebox.showerror("Xato", str(e))
        return
    messagebox.showinfo("OK", f"Titullar yaratildi: {tituls}")

# ------------- DETECTION FRONT (bubbles) Q1..Q35 -------------
def detect_bubbles(img_bgr, coords_df, bubble_window=None, min_black_pixels=None):
    """
    Bubble detection with optional scaled parameters
    coords_df: savol, variant, x, y (already scaled)
    Improved detection with multiple criteria
    
    Args:
        img_bgr: BGR numpy image
        coords_df: DataFrame with scaled coordinates
        bubble_window: Optional scaled bubble window size (defaults to BUBBLE_WINDOW)
        min_black_pixels: Optional scaled minimum black pixels threshold (defaults to MIN_BLACK_PIXELS)
    """

    df = coords_df.copy()
    df.columns = [c.lower() for c in df.columns]
    results = {}
    
    # Use provided scaled parameters or defaults
    win = bubble_window if bubble_window is not None else BUBBLE_WINDOW
    min_pixels = min_black_pixels if min_black_pixels is not None else MIN_BLACK_PIXELS

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    # Use slightly stronger blur for better noise reduction
    gray = cv2.GaussianBlur(gray, (5, 5), 0)

    # Threshold to detect dark marks
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    questions = sorted([q for q in df['savol'].unique() if q <= 35])

    for q in questions:
        subs = df[df['savol'] == q]
        scores = {}

        for _, r in subs.iterrows():
            var = str(r.get('variant', '')).strip().upper()
            if not var:
                continue

            x = int(r['x'])
            y = int(r['y'])

            # Ensure window is properly sized (inclusive on both ends)
            x0 = max(0, x - win // 2)
            x1 = min(th.shape[1], x + win // 2 + 1)  # +1 for inclusive end
            y0 = max(0, y - win // 2)
            y1 = min(th.shape[0], y + win // 2 + 1)  # +1 for inclusive end

            crop = th[y0:y1, x0:x1]

            # Count dark pixels
            black = int(cv2.countNonZero(crop))
            scores[var] = black

        if not scores:
            results[q] = None
            continue

        # Improved detection logic: use multiple criteria
        sorted_vars = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        best_var, best_val = sorted_vars[0]
        
        # Calculate statistics
        all_values = list(scores.values())
        avg = np.mean(all_values)
        std = np.std(all_values) if len(all_values) > 1 else 0
        
        # Get second best value
        second_best_val = sorted_vars[1][1] if len(sorted_vars) > 1 else 0
        
        # Scaled threshold for second best difference (proportional to window size)
        diff_threshold = max(20, int(win * 0.4))  # At least 20 or 40% of window size
        
        # Multiple criteria for detection:
        # 1. Best value must be above absolute minimum threshold
        # 2. Best value must be above average * multiplier
        # 3. Best value should be significantly higher than second best
        condition1 = best_val >= min_pixels
        condition2 = best_val > avg * THRESH_MULT
        condition3 = (best_val - second_best_val) >= diff_threshold
        
        # If best is clearly above average and second best, accept it
        # Also accept if it's well above the average even if second best is close
        if condition1 and (condition2 or (best_val > avg + std and best_val > second_best_val + int(diff_threshold * 0.75))):
            results[q] = best_var
        else:
            results[q] = None

    return results

# ------------- DETECTION BACK (squares) Q36..Q55 -------------
def detect_back_squares(img_bgr, back_coords_df, square_half=None, back_black_threshold=None):
    """
    img_bgr: BGR numpy image
    back_coords_df: df with savol (36..55), x, y (already scaled)
    returns dict {qnum: 0/1} - 1 if square is painted, 0 if not
    
    Args:
        img_bgr: BGR numpy image
        back_coords_df: DataFrame with scaled coordinates
        square_half: Optional scaled half side length for square (defaults to BACK_SQUARE_HALF)
        back_black_threshold: Optional scaled black threshold (defaults to BACK_BLACK_THRESHOLD)
    """
    df = back_coords_df.copy()
    df.columns = [c.lower() for c in df.columns]
    
    # Use provided scaled parameters or defaults
    half_side = square_half if square_half is not None else BACK_SQUARE_HALF
    threshold = back_black_threshold if back_black_threshold is not None else BACK_BLACK_THRESHOLD
    
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5,5), 0)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    results = {}
    for _, r in df.iterrows():
        q = int(r['savol'])
        if q < 36 or q > 55:
            continue
        x = int(r['x'])
        y = int(r['y'])
        
        # Create square region centered at (x, y) with side length = 2 * half_side
        x0 = max(0, x - half_side)
        x1 = min(th.shape[1], x + half_side)
        y0 = max(0, y - half_side)
        y1 = min(th.shape[0], y + half_side)
        
        # Extract square region
        crop = th[y0:y1, x0:x1]
        
        # Count black pixels in the square
        black = int(cv2.countNonZero(crop))
        
        # If painted (black pixels > threshold), write 1, otherwise 0
        val = 1 if black > threshold else 0
        results[q] = val
    return results

# ------------- GRADE SCANS FROM PDF (each page is one titul front image) -------------
def grade_scans_from_pdf():
    pdf_path = filedialog.askopenfilename(title="Skanlangan PDF tanlang (har sahifa — bitta titul)", filetypes=[("PDF files", "*.pdf")])
    if not pdf_path:
        return
    coords = load_pixel_points()
    if coords is None:
        return
    try:
        pages = convert_from_path(pdf_path, dpi=DPI)
    except Exception as e:
        messagebox.showerror("Xato", f"PDF o'qishda xato: {e}")
        return
    students = load_students()
    all_rows = []
    for i, page in enumerate(pages, start=1):
        pil = page.convert("RGB")
        img_width, img_height = pil.size
        # Scale coordinates to match this page's dimensions
        scaled_coords = scale_coordinates(coords, img_width, img_height)
        front_coords = scaled_coords[scaled_coords['savol'] <= 35]
        # Scale detection parameters
        scaled_params = scale_detection_params(img_width, img_height)
        img_np = np.array(pil)[:, :, ::-1]
        codes = decode(pil)
        if not codes:
            print(f"Sahifa {i}: QR topilmadi.")
            continue
        try:
            data = json.loads(codes[0].data.decode("utf-8"))
        except Exception:
            print(f"Sahifa {i}: QR parse hatosi.")
            continue
        test_name = data.get("test")
        sid = str(data.get("id")).replace("TST", "")
        key = load_key(test_name)
        if key is None:
            print(f"{test_name} kaliti topilmadi.")
            continue
        # detect front bubbles with scaled coordinates and parameters
        detected = detect_bubbles(img_np, front_coords,
                                 bubble_window=scaled_params['bubble_window'],
                                 min_black_pixels=scaled_params['min_black_pixels'])
        row = {"Ism Familiya": "", "ID": sid}
        if sid in students["ID"].astype(str).tolist():
            row["Ism Familiya"] = students.loc[students["ID"].astype(str) == sid, "Ism Familiya"].values[0]
        for q in range(1, 36):
            d = detected.get(q)
            try:
                exp = str(key.loc[q-1, "Javob"]).strip().upper()
            except Exception:
                exp = ""
            row[f"Q{q}"] = 1 if (d and exp and d == exp) else 0
        all_rows.append(row)
        print(f"Sahifa {i}: {sid} processed.")
    if not all_rows:
        messagebox.showinfo("Natija", "Hech qanday titul tekshirilmadi.")
        return
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Natijani saqlash (front results)")
    if not save_path:
        return
    pd.DataFrame(all_rows).to_excel(save_path, index=False)
    messagebox.showinfo("OK", f"Natijalar saqlandi: {save_path}")

# ------------- GRADE FRONT FROM IMAGE (JPG/PNG) 1..35 -------------
def grade_front_from_image():
    img_path = filedialog.askopenfilename(title="Old tomon rasm tanlang (JPG/PNG)", filetypes=[("Images", "*.jpg *.jpeg *.png")])
    if not img_path:
        return
    coords = load_pixel_points()
    if coords is None:
        return
    pil = Image.open(img_path).convert("RGB")
    img_width, img_height = pil.size
    # Scale coordinates to match image dimensions
    scaled_coords = scale_coordinates(coords, img_width, img_height)
    front_coords = scaled_coords[scaled_coords['savol'] <= 35]
    # Scale detection parameters
    scaled_params = scale_detection_params(img_width, img_height)
    img_np = np.array(pil)[:, :, ::-1]
    codes = decode(pil)
    if not codes:
        messagebox.showerror("Xato", "QR kod rasmda topilmadi.")
        return
    try:
        data = json.loads(codes[0].data.decode("utf-8"))
    except Exception:
        messagebox.showerror("Xato", "QR parse xato.")
        return
    test_name = data.get("test")
    sid = str(data.get("id")).replace("TST", "")
    key = load_key(test_name)
    if key is None:
        messagebox.showerror("Xato", f"{test_name} kaliti topilmadi.")
        return
    detected = detect_bubbles(img_np, front_coords, 
                              bubble_window=scaled_params['bubble_window'],
                              min_black_pixels=scaled_params['min_black_pixels'])
    students = load_students()
    row = {"Ism Familiya": "", "ID": sid}
    if sid in students["ID"].astype(str).tolist():
        row["Ism Familiya"] = students.loc[students["ID"].astype(str) == sid, "Ism Familiya"].values[0]
    for q in range(1, 36):
        d = detected.get(q)
        try:
            exp = str(key.loc[q-1, "Javob"]).strip().upper()
        except:
            exp = ""
        row[f"Q{q}"] = 1 if (d and exp and d == exp) else 0
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Natijani saqlash (front image)")
    if not save_path:
        return
    pd.DataFrame([row]).to_excel(save_path, index=False)
    messagebox.showinfo("OK", f"Old tomonni tekshirish saqlandi: {save_path}")

# ------------- GRADE MULTIPLE IMAGES (BATCH PROCESSING) 1..35 -------------
def grade_multiple_images():
    """Process multiple answer sheet images at once"""
    img_paths = filedialog.askopenfilenames(
        title="Ko'p sonli javob varaqlarini tanlang (JPG/PNG)", 
        filetypes=[("Images", "*.jpg *.jpeg *.png"), ("All files", "*.*")]
    )
    if not img_paths:
        return
    
    coords = load_pixel_points()
    if coords is None:
        return
    students = load_students()
    
    all_rows = []
    processed_count = 0
    error_count = 0
    errors = []
    
    # Create progress window
    progress_window = tk.Toplevel(root)
    progress_window.title("Jarayon...")
    progress_window.geometry("400x150")
    tk.Label(progress_window, text=f"Jami: {len(img_paths)} ta rasm", font=("Helvetica", 12)).pack(pady=10)
    progress_label = tk.Label(progress_window, text="Tayyorlanmoqda...", font=("Helvetica", 10))
    progress_label.pack(pady=5)
    progress_window.update()
    
    for idx, img_path in enumerate(img_paths, 1):
        try:
            progress_label.config(text=f"Tekshirilmoqda: {idx}/{len(img_paths)} - {os.path.basename(img_path)}")
            progress_window.update()
            
            pil = Image.open(img_path).convert("RGB")
            img_width, img_height = pil.size
            # Scale coordinates to match this image's dimensions
            scaled_coords = scale_coordinates(coords, img_width, img_height)
            front_coords = scaled_coords[scaled_coords['savol'] <= 35]
            # Scale detection parameters
            scaled_params = scale_detection_params(img_width, img_height)
            img_np = np.array(pil)[:, :, ::-1]
            codes = decode(pil)
            
            if not codes:
                error_count += 1
                errors.append(f"{os.path.basename(img_path)}: QR kod topilmadi")
                continue
            
            try:
                data = json.loads(codes[0].data.decode("utf-8"))
            except Exception:
                error_count += 1
                errors.append(f"{os.path.basename(img_path)}: QR parse xato")
                continue
            
            test_name = data.get("test")
            sid = str(data.get("id")).replace("TST", "")
            key = load_key(test_name)
            
            if key is None:
                error_count += 1
                errors.append(f"{os.path.basename(img_path)}: {test_name} kaliti topilmadi")
                continue
            
            # Detect bubbles with scaled coordinates and parameters
            detected = detect_bubbles(img_np, front_coords,
                                     bubble_window=scaled_params['bubble_window'],
                                     min_black_pixels=scaled_params['min_black_pixels'])
            
            # Create result row
            row = {"Ism Familiya": "", "ID": sid, "Test": test_name, "Fayl": os.path.basename(img_path)}
            if sid in students["ID"].astype(str).tolist():
                row["Ism Familiya"] = students.loc[students["ID"].astype(str) == sid, "Ism Familiya"].values[0]
            
            # Grade each question
            for q in range(1, 36):
                d = detected.get(q)
                try:
                    exp = str(key.loc[q-1, "Javob"]).strip().upper()
                except Exception:
                    exp = ""
                row[f"Q{q}"] = 1 if (d and exp and d == exp) else 0
            
            all_rows.append(row)
            processed_count += 1
            
        except Exception as e:
            error_count += 1
            errors.append(f"{os.path.basename(img_path)}: {str(e)}")
            continue
    
    progress_window.destroy()
    
    if not all_rows:
        messagebox.showerror("Xato", f"Hech qanday rasm muvaffaqiyatli tekshirilmadi.\n\nXatolar:\n" + "\n".join(errors[:10]))
        return
    
    # Save results
    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", 
        filetypes=[("Excel", "*.xlsx")], 
        title="Natijani saqlash (ko'p sonli rasmlar)"
    )
    if not save_path:
        return
    
    df_results = pd.DataFrame(all_rows)
    df_results.to_excel(save_path, index=False)
    
    # Show summary
    summary = f"Muvaffaqiyatli tekshirildi: {processed_count}/{len(img_paths)}\n"
    if error_count > 0:
        summary += f"Xatolar: {error_count}\n"
        if errors:
            summary += f"\nXato ro'yxati:\n" + "\n".join(errors[:5])
            if len(errors) > 5:
                summary += f"\n... va yana {len(errors) - 5} ta xato"
    
    messagebox.showinfo("OK", f"Natijalar saqlandi: {save_path}\n\n{summary}")

# ------------- GRADE BACK FROM IMAGE (36..55) -------------
def grade_back_from_image():
    img_path = filedialog.askopenfilename(title="Orqa tomon rasm tanlang (JPG/PNG)", filetypes=[("Images", "*.jpg *.jpeg *.png")])
    if not img_path:
        return
    coords = load_pixel_points()
    if coords is None:
        return
    pil = Image.open(img_path).convert("RGB")
    img_width, img_height = pil.size
    # Scale coordinates to match image dimensions
    scaled_coords = scale_coordinates(coords, img_width, img_height)
    back_coords = scaled_coords[(scaled_coords['savol'] >= 36) & (scaled_coords['savol'] <= 55)]
    if back_coords.empty:
        messagebox.showerror("Xato", "Orqa uchun koordinatalar topilmadi.")
        return
    # Scale detection parameters
    scaled_params = scale_detection_params(img_width, img_height)
    img_np = np.array(pil)[:, :, ::-1]
    results = detect_back_squares(img_np, back_coords,
                                   square_half=scaled_params['back_square_half'],
                                   back_black_threshold=scaled_params['back_black_threshold'])
    row = {}
    for q in range(36, 56):
        row[f"Q{q}"] = results.get(q, 0)
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Orqa natijalarini saqlash")
    if not save_path:
        return
    pd.DataFrame([row]).to_excel(save_path, index=False)
    messagebox.showinfo("OK", f"Orqa tomonni tekshirish saqlandi: {save_path}")

# ------------- GRADE BACK FROM PDF PAGE 2 -------------
def grade_back_from_pdf_page():
    pdf_path = filedialog.askopenfilename(title="PDF tanlang (orqa page ichidan)", filetypes=[("PDF", "*.pdf")])
    if not pdf_path:
        return
    coords = load_pixel_points()
    if coords is None:
        return
    try:
        pages = convert_from_path(pdf_path, dpi=DPI)
    except Exception as e:
        messagebox.showerror("Xato", f"PDF o'qishda xato: {e}")
        return
    if len(pages) < 2:
        messagebox.showerror("Xato", "PDF ichida orqa sahifa topilmadi.")
        return
    pil = pages[1].convert("RGB")
    img_width, img_height = pil.size
    # Scale coordinates to match page dimensions
    scaled_coords = scale_coordinates(coords, img_width, img_height)
    back_coords = scaled_coords[(scaled_coords['savol'] >= 36) & (scaled_coords['savol'] <= 55)]
    if back_coords.empty:
        messagebox.showerror("Xato", "Orqa uchun koordinatalar topilmadi.")
        return
    # Scale detection parameters
    scaled_params = scale_detection_params(img_width, img_height)
    img_np = np.array(pil)[:, :, ::-1]
    results = detect_back_squares(img_np, back_coords,
                                  square_half=scaled_params['back_square_half'],
                                  back_black_threshold=scaled_params['back_black_threshold'])
    row = {}
    for q in range(36, 56):
        row[f"Q{q}"] = results.get(q, 0)
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], title="Orqa natijalarini saqlash")
    if not save_path:
        return
    pd.DataFrame([row]).to_excel(save_path, index=False)
    messagebox.showinfo("OK", f"Orqa sahifa tekshirildi va saqlandi: {save_path}")

# ------------- RASCH MODEL (55 items) -------------
def run_rasch_model():
    input_file = filedialog.askopenfilename(title="Rasch uchun results.xlsx ni tanlang (Q1..Q55)", filetypes=[("Excel",".xlsx")])
    if not input_file:
        return
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        messagebox.showerror("Xato", f"Faylni o'qishda xato: {e}")
        return
    name_col = df.columns[0]
    id_col = df.columns[1] if len(df.columns) > 1 else None
    question_cols = [c for c in df.columns if str(c).upper().startswith("Q")]
    if len(question_cols) != 55:
        messagebox.showwarning("Ogohlantirish", f"{len(question_cols)} ta savol topildi. 55 bo'lishi kerak. Davom etiladi.")
    names = df[name_col]
    ids = df[id_col] if id_col is not None else pd.Series([""] * len(df))
    answers = df[question_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
    n_students, n_items = answers.shape
    weights = np.ones(n_items)
    if n_items > 35:
        weights[35:] = 1.5
    def neg_ll(params):
        theta = params[:n_students]
        beta = params[n_students:]
        P = 1 / (1 + np.exp(-(theta.reshape(-1,1) - beta.reshape(1,-1))))
        P = np.clip(P, 1e-9, 1-1e-9)
        ll = np.sum(weights * (answers.values * np.log(P) + (1 - answers.values) * np.log(1-P)))
        return -ll
    params0 = np.zeros(n_students + n_items)
    res = minimize(neg_ll, params0, method="L-BFGS-B", options={"maxiter":500})
    params = res.x
    theta = params[:n_students]
    beta = params[n_students:]
    beta -= np.mean(beta)
    Z = (theta - np.mean(theta)) / np.std(theta)
    T = 50 + 10 * Z
    T = np.clip(T, 10, 90)
    def baho(t):
        if t >= 70: return "A+"
        if t >= 65: return "A"
        if t >= 60: return "B+"
        if t >= 55: return "B"
        if t >= 50: return "C+"
        if t >= 45: return "C"
        return "Sertifikat berilmadi"
    grades = [baho(t) for t in T]
    raw_scores = answers.sum(axis=1)
    results = pd.DataFrame({
        "Ism Familiya": names,
        "ID": ids,
        "To‘g‘ri javoblar soni": raw_scores,
        "Qobiliyat (θ)": np.round(theta,3),
        "Z ball": np.round(Z,3),
        "T ball": np.round(T,2),
        "Baholash": grades
    })
    diff = pd.DataFrame({
        "Savol": question_cols,
        "Qiyinlik (β)": np.round(beta,3)
    })
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")], title="Rasch natijalarini saqlash")
    if not save_path:
        return
    with pd.ExcelWriter(save_path) as writer:
        results.to_excel(writer, sheet_name="Natijalar", index=False)
        diff.to_excel(writer, sheet_name="Savol Qiyinliklari", index=False)
    wb = load_workbook(save_path)
    ws = wb["Natijalar"]
    colors = {"A+": "000080", "A":"3399FF", "B+":"00FF00", "B":"FFFF00", "C+":"FF9900", "C":"FFCC00", "Sertifikat berilmadi":"FF0000"}
    grade_col = list(results.columns).index("Baholash") + 1
    def col_letter(n):
        s = ""
        while n > 0:
            n, r = divmod(n-1, 26)
            s = chr(65+r) + s
        return s
    grade_letter = col_letter(grade_col)
    for r in range(2, ws.max_row + 1):
        grade_value = ws[f"{grade_letter}{r}"].value
        col = colors.get(grade_value)
        if col:
            fill = PatternFill(start_color=col, end_color=col, fill_type="solid")
            for c in range(1, ws.max_column + 1):
                ws[f"{col_letter(c)}{r}"].fill = fill
    wb.save(save_path)
    messagebox.showinfo("OK", f"Rasch modeli bajarildi va saqlandi: {save_path}")

# ------------- RUN RASH.PY SCRIPT -------------
def run_rash_script():
    """Run the rash.py script with file selection"""
    # Ask for input file
    input_file = filedialog.askopenfilename(
        title="Rash modeli uchun input fayl tanlang (result_1.xlsx)", 
        filetypes=[("Excel", ".xlsx"), ("All files", "*.*")],
        initialdir=BASE_DIR
    )
    if not input_file:
        return
    
    # Ask for output file location
    output_file = filedialog.asksaveasfilename(
        title="Natijani saqlash", 
        defaultextension=".xlsx",
        filetypes=[("Excel", ".xlsx")],
        initialdir=BASE_DIR,
        initialfile="rasch_55_weight_chart.xlsx"
    )
    if not output_file:
        return
    
    try:
        # Read and process data (same logic as rash.py)
        df = pd.read_excel(input_file)
        
        name_col = df.columns[0]
        id_col = df.columns[1] if len(df.columns) > 1 else None
        question_cols = [c for c in df.columns if str(c).upper().startswith("Q")]
        
        if len(question_cols) != 55:
            messagebox.showwarning("Ogohlantirish", f"{len(question_cols)} ta savol topildi. 55 bo'lishi kerak. Davom etiladi.")
        
        names = df[name_col]
        ids = df[id_col] if id_col is not None else pd.Series([""] * len(df))
        answers = df[question_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)
        
        n_students, n_items = answers.shape
        
        weights = np.ones(n_items)
        weights[35:] = 1.5  # 36–55 savollar og'irroq
        
        def neg_ll(params):
            theta = params[:n_students]
            beta = params[n_students:]
            P = 1 / (1 + np.exp(-(theta.reshape(-1,1) - beta.reshape(1,-1))))
            P = np.clip(P, 1e-9, 1-1e-9)
            ll = np.sum(weights * (answers.values * np.log(P) + (1-answers.values) * np.log(1-P)))
            return -ll
        
        params_init = np.zeros(n_students + n_items)
        res = minimize(neg_ll, params_init, method="L-BFGS-B", options={"maxiter":500})
        params = res.x
        
        theta = params[:n_students]
        beta = params[n_students:]
        beta -= np.mean(beta)
        
        Z = (theta - np.mean(theta)) / np.std(theta)
        T = np.clip(50 + 10 * Z, 10, 90)
        
        def baho(t):
            if t >= 70: return "A+"
            if t >= 65: return "A"
            if t >= 60: return "B+"
            if t >= 55: return "B"
            if t >= 50: return "C+"
            if t >= 45: return "C"
            return "Sertifikat berilmadi"
        
        grades = [baho(t) for t in T]
        raw_scores = answers.sum(axis=1)
        
        results = pd.DataFrame({
            "Ism Familiya": names,
            "ID": ids,
            "To'g'ri javoblar soni": raw_scores,
            "Qobiliyat (θ)": np.round(theta,3),
            "Z ball": np.round(Z,3),
            "T ball": np.round(T,2),
            "Baholash": grades
        })
        
        # Sort by T descending
        results = results.sort_values(by="T ball", ascending=False)
        
        diff = pd.DataFrame({
            "Savol": question_cols,
            "Qiyinlik (β)": np.round(beta,3)
        })
        
        with pd.ExcelWriter(output_file) as writer:
            results.to_excel(writer, sheet_name="Natijalar", index=False)
            diff.to_excel(writer, sheet_name="Savol Qiyinliklari", index=False)
        
        # Apply colors and create chart (same as rash.py)
        wb = load_workbook(output_file)
        ws = wb["Natijalar"]
        ws_diff = wb["Savol Qiyinliklari"]
        
        colors = {
            "A+": "000080",
            "A": "3399FF",
            "B+": "00FF00",
            "B": "FFFF00",
            "C+": "FF9900",
            "C": "FFCC00",
            "Sertifikat berilmadi": "FF0000"
        }
        
        grade_col = list(results.columns).index("Baholash") + 1
        grade_letter = chr(ord("A") + grade_col - 1)
        
        for r in range(2, ws.max_row + 1):
            grade_value = ws[f"{grade_letter}{r}"].value
            color = colors.get(grade_value)
            if color:
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for c in range(1, ws.max_column + 1):
                    ws[f"{chr(64+c)}{r}"].fill = fill
        
        # Create bar chart (difficulty diagram)
        chart = BarChart()
        chart.title = "Savollar qiyinligi (β)"
        chart.x_axis.title = "Savollar"
        chart.y_axis.title = "Qiyinlik"
        
        data = Reference(ws_diff, min_col=2, min_row=1, max_row=ws_diff.max_row)
        cats = Reference(ws_diff, min_col=1, min_row=2, max_row=ws_diff.max_row)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_diff.add_chart(chart, "E2")
        
        wb.save(output_file)
        messagebox.showinfo("Muvaffaqiyatli", f"🎉 Tayyor! Diagramma qo'shildi va natijalar T bo'yicha tartiblandi:\n{output_file}")
        
    except Exception as e:
        messagebox.showerror("Xato", f"Rash modelini bajarishda xato yuz berdi:\n{str(e)}")
        import traceback
        traceback.print_exc()

#########################
#  GUI ENTRY POINT
#########################

if __name__ == "__main__":
    # ------------- GUI -------------
    # Check keyword verification before starting GUI
    if not check_keyword_verification():
        sys.exit(0)

    root = tk.Tk()
    root.title("Rash Test Manager v2 — Full")
    root.geometry("820x700")

    tk.Label(root, text="Rash Test Manager v2", font=("Helvetica", 20, "bold")).pack(pady=10)
    frm = tk.Frame(root)
    frm.pack(pady=6)

    # GUI actions
    def add_student_action():
        name = simpledialog.askstring("O'quvchi qo'shish", "Ism Familiya:")
        if not name: return
        sid = simpledialog.askstring("O'quvchi qo'shish", "ID:")
        if not sid: return
        df = load_students()
        if str(sid) in df["ID"].astype(str).tolist():
            messagebox.showerror("Xato", "Bu ID allaqachon mavjud.")
            return
        df = pd.concat([df, pd.DataFrame([{"Ism Familiya": name, "ID": sid}])], ignore_index=True)
        save_students(df)
        messagebox.showinfo("OK", f"{name} qo‘shildi.")

    def create_new_test_action():
        test_name = simpledialog.askstring("Yangi test", "Test nomini kiriting:")
        if not test_name: return
        ensure_test_dirs(test_name)
        create_key_if_missing(test_name)
        messagebox.showinfo("OK", f"Test '{test_name}' yaratildi.")

    def open_key_file():
        p = filedialog.askopenfilename(initialdir=KEYS_DIR, title="Kalit tanlang", filetypes=[("Excel",".xlsx")])
        if p:
            os.system(f'open "{p}"' if os.name == 'posix' else f'start "" "{p}"')

    tk.Button(frm, text="🧑‍🎓 O'quvchi qo'shish", width=30, command=add_student_action).grid(row=0, column=0, padx=6, pady=6)
    tk.Button(frm, text="➕ Yangi test yaratish", width=30, command=create_new_test_action).grid(row=0, column=1, padx=6, pady=6)
    tk.Button(frm, text="🔑 Kalitni ochish", width=30, command=open_key_file).grid(row=1, column=0, padx=6, pady=6)
    tk.Button(frm, text="🖨️ Titul yaratish (OLD+ORQA PDF)", width=30, command=generate_tituls).grid(row=1, column=1, padx=6, pady=6)


    tk.Button(frm, text="🖼 Old tomonni tekshirish (JPG/PNG -> Q1..Q35)", width=62, command=grade_front_from_image).grid(row=3, column=0, columnspan=2, pady=6)
    tk.Button(frm, text="📁 Ko'p sonli rasmlarni tekshirish (JPG/PNG -> Q1..Q35)", width=62, command=grade_multiple_images, bg="#4CAF50", fg="black").grid(row=4, column=0, columnspan=2, pady=6)
    tk.Button(frm, text="🖼 Orqa tomonni tekshirish (JPG/PNG -> Q36..Q55)", width=62, command=grade_back_from_image).grid(row=5, column=0, columnspan=2, pady=6)


    tk.Button(frm, text="📈 Rasch modeli (55 savol)", width=62, command=run_rasch_model).grid(row=7, column=0, columnspan=2, pady=6)
    tk.Button(frm, text="📊 Rash modeli (rash.py - diagramma bilan)", width=62, command=run_rash_script, bg="#9C27B0", fg="white").grid(row=8, column=0, columnspan=2, pady=6)

    tk.Label(root, text="Shablon: 12.pdf (~/Desktop/titul_system/12.pdf yoki /mnt/data/12.pdf)", fg="gray").pack(pady=6)
    tk.Button(root, text="❌ Chiqish", bg="#f44336", fg="white", command=root.destroy).pack(pady=10)

    root.mainloop()