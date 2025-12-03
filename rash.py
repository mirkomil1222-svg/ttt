import pandas as pd
import numpy as np
from scipy.optimize import minimize
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

INPUT_FILE = "result_1.xlsx"
OUTPUT_FILE = "rasch_55_weight_chart.xlsx"

df = pd.read_excel(INPUT_FILE)

name_col = df.columns[0]
id_col = df.columns[1]
question_cols = [c for c in df.columns if str(c).upper().startswith("Q")]

names = df[name_col]
ids = df[id_col]
answers = df[question_cols].apply(pd.to_numeric, errors="coerce").fillna(0).astype(int)

n_students, n_items = answers.shape

weights = np.ones(n_items)
weights[35:] = 1.5  # 36–55 savollar og‘irroq

def neg_ll(params):
    theta = params[:n_students]
    beta = params[n_students:]
    P = 1 / (1 + np.exp(-(theta.reshape(-1,1) - beta.reshape(1,-1))))
    P = np.clip(P, 1e-9, 1-1e-9)
    ll = np.sum(weights * (answers.values * np.log(P) + (1-answers.values) * np.log(1-P)))
    return -ll

params_init = np.zeros(n_students + n_items)
res = minimize(neg_ll, params_init, method="L-BFGS-B", options={"maxiter":500})
params = res.x

theta = params[:n_students]
beta = params[n_students:]
beta -= np.mean(beta)

Z = (theta - np.mean(theta)) / np.std(theta)
T = np.clip(50 + 10 * Z, 10, 90)

def baho(t):
    if t >= 70: return "A+"
    if t >= 65: return "A"
    if t >= 60: return "B+"
    if t >= 55: return "B"
    if t >= 50: return "C+"
    if t >= 45: return "C"
    return "Sertifikat berilmadi"

grades = [baho(t) for t in T]
raw_scores = answers.sum(axis=1)

results = pd.DataFrame({
    "Ism Familiya": names,
    "ID": ids,
    "To‘g‘ri javoblar soni": raw_scores,
    "Qobiliyat (θ)": np.round(theta,3),
    "Z ball": np.round(Z,3),
    "T ball": np.round(T,2),
    "Baholash": grades
})

# Sort by T descending
results = results.sort_values(by="T ball", ascending=False)

diff = pd.DataFrame({
    "Savol": question_cols,
    "Qiyinlik (β)": np.round(beta,3)
})

with pd.ExcelWriter(OUTPUT_FILE) as writer:
    results.to_excel(writer, sheet_name="Natijalar", index=False)
    diff.to_excel(writer, sheet_name="Savol Qiyinliklari", index=False)

wb = load_workbook(OUTPUT_FILE)
ws = wb["Natijalar"]
ws_diff = wb["Savol Qiyinliklari"]

colors = {
    "A+": "000080",
    "A": "3399FF",
    "B+": "00FF00",
    "B": "FFFF00",
    "C+": "FF9900",
    "C": "FFCC00",
    "Sertifikat berilmadi": "FF0000"
}

grade_col = list(results.columns).index("Baholash") + 1
grade_letter = chr(ord("A") + grade_col - 1)

for r in range(2, ws.max_row + 1):
    grade_value = ws[f"{grade_letter}{r}"].value
    color = colors.get(grade_value)
    if color:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for c in range(1, ws.max_column + 1):
            ws[f"{chr(64+c)}{r}"].fill = fill

# === Create bar chart (difficulty diagram) ===
chart = BarChart()
chart.title = "Savollar qiyinligi (β)"
chart.x_axis.title = "Savollar"
chart.y_axis.title = "Qiyinlik"

data = Reference(ws_diff, min_col=2, min_row=1, max_row=ws_diff.max_row)
cats = Reference(ws_diff, min_col=1, min_row=2, max_row=ws_diff.max_row)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws_diff.add_chart(chart, "E2")

wb.save(OUTPUT_FILE)
print("🎉 Tayyor! Diagramma qo‘shildi va natijalar T bo‘yicha tartiblandi:", OUTPUT_FILE)